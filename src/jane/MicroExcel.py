
'''
Created on 24Jul.,2018

@author: pwang
'''
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
from openpyxl.styles import Font, Fill
from openpyxl.utils.dataframe import dataframe_to_rows 
import NormalizeRawData as nrd
from NormalizeRawData import COLINDEX as ci
import pandas as pd
import numpy as np
import os
from Template import Template




class MicroExcel(object):

    def __init__(self):
        self.CONFIG_SHEET_ENTITIES = "entities"
        self.CONFIG_SHEET_ACCOUNTS = "accounts"
        self.CONFIG_SHEET_TEMPLATES='templates'
        self.sheetName="interface1"
        self.CONFIG_ENTITIES_FILENAME = 'fileName'
        self.CONFIG_ENTITIES_BUDGET = 'BUDGET'
        self.CONFIG_ENTITIES_ACTUAL = 'ACTUAL'
        self.CONFIG_ENTITIES_PRIOR = 'PRIOR'
      
       

        
    def loadSheet(self,path,fileName):
        budgetDf=pd.DataFrame()
        actualDf=pd.DataFrame()
        priorDf=pd.DataFrame()
        configFile=path+fileName
        #print("configFile=="+configFile)
        self.wb=openpyxl.load_workbook(configFile)
        self.entities=self.wb[self.CONFIG_SHEET_ENTITIES]
        config_cells=self.entities.__getitem__("A3:F200")
        accountSheet=self.wb[self.CONFIG_SHEET_ACCOUNTS]
        self.templates=self.wb[self.CONFIG_SHEET_TEMPLATES]
        config_account_cells=accountSheet.__getitem__("A2:A200")
        self.accountList=nrd.leanupExcelAccountTupleCells(config_account_cells)
        for i in range(len(config_cells)):
            
            entitiesFileName=config_cells[i][0].value
            if pd.isnull(entitiesFileName)==True:
                break
                
            if(pd.isnull(entitiesFileName)!=True):
                entitiesFileName=path+entitiesFileName
                priorExist=config_cells[i][1].value
                budgetExist=config_cells[i][2].value
                actualExist=config_cells[i][3].value
                b19Exist=config_cells[i][4].value
                sheetRange=config_cells[i][5].value
                wb2=openpyxl.load_workbook(entitiesFileName)
                if(budgetExist!=''):
                    budgetSheets=wb2[budgetExist]
    
                    cells=budgetSheets.__getitem__(sheetRange)
                    budgetDf1=nrd.normalizeExcel(ci.INDEX_F18,cells)
                    #budgetDf=nrd.reduceRepositoryByAccounts(budgetDf, self.accountList)
                    
                    budgetDf=budgetDf.append(budgetDf1)
               
    
                if(actualExist!=''):   
                    actualSheet=wb2[actualExist]
                   
                    cells=actualSheet.__getitem__(sheetRange)
                    actualDf1=nrd.normalizeExcel(ci.INDEX_B18,cells)
                    #actualDf=nrd.reduceRepositoryByAccounts(actualDf, self.accountList)
                    
                    actualDf=actualDf.append(actualDf1)
              
                
                if(priorExist!=''):
                    priorSheet=wb2[priorExist]
                    cells=priorSheet.__getitem__(sheetRange)
                    priorDf1=nrd.normalizeExcel(ci.INDEX_A17,cells)
                    #priorDf=nrd.reduceRepositoryByAccounts(priorDf, self.accountList)
                    priorDf=priorDf.append(priorDf1)
               
               
                if(pd.isnull(b19Exist)!=True):
                    b19sheet=wb2[b19Exist]
                    cells=b19sheet.__getitem__(sheetRange)
                    b19Df=nrd.normalizeExcel(ci.INDEX_B19,cells)
                    

        #b19 is from different files. if it is empty, then set it to zero 
        if(pd.isnull(b19Exist)!=True):
            b19=nrd.reduceRepositoryByAccounts(b19Df, self.accountList)
        else:
            b19=nrd.generateB19(budgetDf)
            
        budgetDf=nrd.reduceRepositoryByAccounts(budgetDf, self.accountList)
        actualDf=nrd.reduceRepositoryByAccounts(actualDf, self.accountList)
        priorDf=nrd.reduceRepositoryByAccounts(priorDf, self.accountList) 
        b19=nrd.reduceRepositoryByAccounts(b19, self.accountList)      
        self.repos= self.generateRepos(priorDf,actualDf,budgetDf,b19)                       

            
    
    
    def loadTemplateWorkSheet(self,sheetName):
        presentSheets =self.wb[sheetName]
        pCells=presentSheets.__getitem__("A2:F150")
        df=pd.DataFrame(columns=ci.CONFIG_COLUMNS)
        for i in range(0,len(pCells)):
            if(pd.isnull(pCells[i][0].value)!=True):
                row=[pCells[i][0].value,pCells[i][1].value,pCells[i][2].value,pCells[i][3].value,pCells[i][4].value,pCells[i][5].value]
                df.loc[len(df)]=row
        return df
    
    

    
    def generateRepos(self,priorDf,actualDf,budgetDf,b19Df):
        repos=pd.DataFrame()
        repos=repos.append(priorDf)
        repos=repos.append(actualDf)
        repos=repos.append(budgetDf)
        repos=repos.append(b19Df)
        # ensure repos don't have nan value in order to exception 
        repos=repos.replace(np.NaN,0)
        repos=repos.append(nrd.generatef18a17(repos))
        repos=repos.append(nrd.generatef18b18(repos))
        repos=repos.append(nrd.generateb19f18(repos))
        
        return repos
    
        
    def getSheet(self,outputFile):
        if(os.path.isfile(outputFile)):
            self.workbook=openpyxl.load_workbook(outputFile) #load old booksheet
        else:
            self.workbook=openpyxl.Workbook()#create new workbook
            
            
    def printTemplates(self,abPath,outputData,template):
        templateDFs=pd.DataFrame(self.templates.values) # convert worksheets to dataFrame structure
        for i in range(1,len(templateDFs)): #loop all templates
            path=templateDFs.iloc[i][0]
            if(pd.isnull(path)):
                fileName=abPath+templateDFs.iloc[i][1]
            else:
                fileName=path+templateDFs.iloc[i][1]
            sheet=templateDFs.iloc[i][2]
            
            self.writeToSheet(fileName, sheet,outputData, template)
                      
                       
    def writeToSheet(self,outputFile,sheetName,outputData,templateWorkSheet):
        self.getSheet(outputFile)
        sheet=self.workbook.create_sheet(title=sheetName)
        
        #print head months
        row=1
        col=1
        sheet.cell(row,col,ci.INDEX_NAME)
        col=col+1
        for mon in ci.COLUMNS_WITHOUT_INDEXES:
            
            end_col=col+6
            sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
            sheet.cell(1,col,mon)
            col=end_col+1
           
        # print sub Head as A17,B18...
        row=2
        col=1
        for column in outputData.columns.tolist():
            sheet.cell(row,col,column)
            col=col+1
        # a blank line 
        # print data
        row=4
        templateClass=Template()
        for templateRow in templateWorkSheet.iterrows():

            d=outputData.query(ci.INDEX_NAME+"=='"+templateRow[1][ci.CONFIG_ACCOUNT]+"'")
            if(~d.empty and len(d)>0):
                # templateWorkSheet row is not percentages of other row
                sheet.cell(row,1,d.iloc[0][0])
                for j in range(1,len(d.columns)):
                    col=j+1
                    if(d.columns[j] in ci.INDEX_PERCENTAGES):
                        sheet.cell(row,col,"{:1.0%}".format(d.iloc[0][j]))
                    else:
                        sheet.cell(row,col,"{:,.0f}".format(d.iloc[0][j]))
            else:
                
                if(templateRow[1][ci.CONFIG_LEVEL]==10):
                    percent=templateClass.getPercentages(templateRow[1][ci.CONFIG_ACCOUNT], templateRow[1][ci.CONFIG_PRECENTAGES], templateRow[1][ci.CONFIG_DENOMINATOR], outputData)
                    sheet.cell(row,1,percent.iloc[0][0])
                    for j in range(1,len(percent.columns)):
                        col=j+1
                        #format to %

                        sheet.cell(row,col,"{:1.0%}".format(percent.iloc[0][j]))
                        
                else:
                    sheet.cell(row,1,templateRow[1][ci.CONFIG_ACCOUNT])
#             if(templateRow[1][ci.CONFIG_LEVEL!='1']):
#                 rowA=sheet.row_dimensions[row]
#                 rowA.font=Font(b="border")
            row=row+1 

        #self.wb.save(self.fileName)

        self.workbook.save(outputFile)
if __name__ == "__main__":
    ie=MicroExcel()
    ie.loadSheet(ie.fileName)
    print(ie.repos)
    



